import numpy as np
import pandas as pd
import logging
from typing import Any, Dict, List, Optional, Tuple, Union
from pathlib import Path
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import warnings

logger = logging.getLogger(__name__)


class StabilityError(Exception):
    pass


class StabilityAnalyzer:
    """
    Analyzer for calculating Population Stability Index (PSI) and Characteristic Stability Index (CSI)
    between datasets, with comprehensive reporting and visualization capabilities.
    """
    
    # PSI/CSI interpretation thresholds
    PSI_THRESHOLDS = {
        'stable': 0.1,
        'slight_shift': 0.2,
        'significant_shift': float('inf')
    }
    
    def __init__(self, bins: int = 10, min_bin_size: float = 0.05):
        """
        Initialize StabilityAnalyzer.
        
        Args:
            bins: Number of bins for discretization (default: 10)
            min_bin_size: Minimum bin size as fraction of total samples (default: 0.05)
        """
        self.bins = bins
        self.min_bin_size = min_bin_size
        self.bin_edges_ = {}
        self.reference_distributions_ = {}
        self.stability_results_ = {}
        
    def fit_reference(self, reference_df: pd.DataFrame, 
                     columns: Optional[List[str]] = None) -> 'StabilityAnalyzer':
        """
        Fit the analyzer to reference dataset (typically training data).
        
        Args:
            reference_df: Reference DataFrame
            columns: Columns to analyze (None for all numeric columns)
            
        Returns:
            self
        """
        logger.info("Fitting StabilityAnalyzer to reference data")
        
        if columns is None:
            columns = reference_df.select_dtypes(include=[np.number]).columns.tolist()
        
        self.columns_ = columns
        
        for column in columns:
            if column not in reference_df.columns:
                logger.warning(f"Column {column} not found in reference data")
                continue
            
            try:
                # Create bins and calculate reference distribution
                bin_edges, ref_distribution = self._calculate_distribution(
                    reference_df[column], column, fit_bins=True
                )
                
                self.bin_edges_[column] = bin_edges
                self.reference_distributions_[column] = ref_distribution
                
                logger.debug(f"Fitted reference distribution for {column}")
                
            except Exception as e:
                logger.error(f"Failed to fit reference for column {column}: {str(e)}")
                continue
        
        logger.info(f"StabilityAnalyzer fitted for {len(self.bin_edges_)} columns")
        return self
    
    def calculate_stability(self, current_df: pd.DataFrame, 
                          dataset_name: str = "current") -> Dict[str, Dict[str, Any]]:
        """
        Calculate PSI/CSI between reference and current datasets.
        
        Args:
            current_df: Current DataFrame to compare against reference
            dataset_name: Name identifier for the current dataset
            
        Returns:
            Dictionary with stability metrics for each column
        """
        logger.info(f"Calculating stability metrics for dataset: {dataset_name}")
        
        if not self.bin_edges_:
            raise StabilityError("Analyzer not fitted. Call fit_reference() first.")
        
        results = {}
        
        for column in self.columns_:
            if column not in current_df.columns:
                logger.warning(f"Column {column} not found in current data")
                continue
            
            try:
                # Calculate current distribution using fitted bins
                _, current_distribution = self._calculate_distribution(
                    current_df[column], column, fit_bins=False
                )
                
                # Calculate PSI
                psi_value = self._calculate_psi(
                    self.reference_distributions_[column],
                    current_distribution
                )
                
                # Determine stability status
                stability_status = self._interpret_psi(psi_value)
                
                # Store detailed results
                results[column] = {
                    'psi': psi_value,
                    'stability_status': stability_status,
                    'reference_distribution': self.reference_distributions_[column],
                    'current_distribution': current_distribution,
                    'bin_edges': self.bin_edges_[column],
                    'dataset_name': dataset_name
                }
                
                logger.debug(f"PSI for {column}: {psi_value:.4f} ({stability_status})")
                
            except Exception as e:
                logger.error(f"Failed to calculate PSI for column {column}: {str(e)}")
                results[column] = {
                    'psi': np.nan,
                    'stability_status': 'error',
                    'error': str(e),
                    'dataset_name': dataset_name
                }
        
        # Store results
        self.stability_results_[dataset_name] = results
        
        logger.info(f"Stability analysis completed for {len(results)} columns")
        return results
    
    def compare_multiple_datasets(self, datasets: Dict[str, pd.DataFrame]) -> Dict[str, Dict[str, Dict[str, Any]]]:
        """
        Compare multiple datasets against the reference.
        
        Args:
            datasets: Dictionary of {dataset_name: DataFrame} pairs
            
        Returns:
            Nested dictionary with stability results for all datasets
        """
        logger.info(f"Comparing {len(datasets)} datasets against reference")
        
        all_results = {}
        
        for dataset_name, df in datasets.items():
            try:
                results = self.calculate_stability(df, dataset_name)
                all_results[dataset_name] = results
            except Exception as e:
                logger.error(f"Failed to analyze dataset {dataset_name}: {str(e)}")
                all_results[dataset_name] = {'error': str(e)}
        
        return all_results
    
    def get_unstable_variables(self, dataset_name: str = None, 
                             threshold: float = None) -> List[str]:
        """
        Get list of variables that exceed stability threshold.
        
        Args:
            dataset_name: Specific dataset to check (None for all)
            threshold: PSI threshold (None for default 'slight_shift' threshold)
            
        Returns:
            List of unstable variable names
        """
        if threshold is None:
            threshold = self.PSI_THRESHOLDS['slight_shift']
        
        unstable_vars = []
        
        if dataset_name:
            datasets_to_check = [dataset_name]
        else:
            datasets_to_check = list(self.stability_results_.keys())
        
        for ds_name in datasets_to_check:
            if ds_name not in self.stability_results_:
                continue
                
            for column, result in self.stability_results_[ds_name].items():
                psi_value = result.get('psi', np.nan)
                if not np.isnan(psi_value) and psi_value > threshold:
                    unstable_vars.append(f"{ds_name}::{column}")
        
        return unstable_vars
    
    def generate_summary_report(self) -> pd.DataFrame:
        """
        Generate summary report of stability metrics across all datasets.
        
        Returns:
            DataFrame with summary statistics
        """
        logger.info("Generating stability summary report")
        
        if not self.stability_results_:
            raise StabilityError("No stability results available. Run calculate_stability() first.")
        
        summary_data = []
        
        for dataset_name, results in self.stability_results_.items():
            for column, result in results.items():
                if 'error' in result:
                    continue
                    
                summary_data.append({
                    'dataset': dataset_name,
                    'variable': column,
                    'psi': result.get('psi', np.nan),
                    'stability_status': result.get('stability_status', 'unknown')
                })
        
        summary_df = pd.DataFrame(summary_data)
        
        if not summary_df.empty:
            # Add summary statistics
            summary_stats = summary_df.groupby(['dataset']).agg({
                'psi': ['count', 'mean', 'std', 'min', 'max'],
                'stability_status': lambda x: (x == 'stable').sum()
            }).round(4)
            
            summary_stats.columns = ['total_vars', 'mean_psi', 'std_psi', 'min_psi', 'max_psi', 'stable_vars']
            summary_stats['unstable_vars'] = summary_stats['total_vars'] - summary_stats['stable_vars']
            summary_stats['stability_rate'] = (summary_stats['stable_vars'] / summary_stats['total_vars']).round(4)
            
            logger.info(f"Summary report generated for {len(summary_data)} variable-dataset combinations")
        
        return summary_df
    
    def create_visualizations(self, output_dir: str = "./stability_plots", 
                            columns: Optional[List[str]] = None,
                            figsize: Tuple[int, int] = (12, 8)) -> List[str]:
        """
        Create visualizations for stability analysis.
        
        Args:
            output_dir: Directory to save plots
            columns: Specific columns to plot (None for all)
            figsize: Figure size for plots
            
        Returns:
            List of generated plot file paths
        """
        logger.info("Creating stability visualizations")
        
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        plot_files = []
        
        if columns is None:
            columns = self.columns_
        
        # 1. PSI Summary Plot
        summary_plot_path = self._create_psi_summary_plot(output_path, figsize)
        if summary_plot_path:
            plot_files.append(summary_plot_path)
        
        # 2. Distribution Comparison Plots
        for column in columns:
            if column in self.bin_edges_:
                dist_plot_path = self._create_distribution_plot(column, output_path, figsize)
                if dist_plot_path:
                    plot_files.append(dist_plot_path)
        
        # 3. PSI Heatmap
        heatmap_path = self._create_psi_heatmap(output_path, figsize)
        if heatmap_path:
            plot_files.append(heatmap_path)
        
        logger.info(f"Created {len(plot_files)} visualization plots")
        return plot_files
    
    def export_to_excel(self, output_path: str, include_charts: bool = True) -> str:
        """
        Export stability analysis results to Excel with formatting and charts.
        
        Args:
            output_path: Path for Excel file
            include_charts: Whether to include charts in Excel
            
        Returns:
            Path to generated Excel file
        """
        logger.info(f"Exporting stability results to Excel: {output_path}")
        
        if not self.stability_results_:
            raise StabilityError("No stability results available. Run calculate_stability() first.")
        
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create summary sheet
        self._create_summary_sheet(wb)
        
        # Create detailed sheets for each dataset
        for dataset_name in self.stability_results_.keys():
            self._create_dataset_sheet(wb, dataset_name)
        
        # Create distributions sheet
        self._create_distributions_sheet(wb)
        
        # Add charts if requested
        if include_charts:
            self._add_excel_charts(wb)
        
        # Save workbook
        wb.save(output_file)
        
        logger.info(f"Excel report saved to: {output_file}")
        return str(output_file)
    
    def _calculate_distribution(self, series: pd.Series, column: str, 
                              fit_bins: bool = True) -> Tuple[np.ndarray, np.ndarray]:
        """Calculate distribution for a series."""
        series_clean = series.dropna()
        
        if len(series_clean) == 0:
            raise StabilityError(f"No non-null values found in column {column}")
        
        if fit_bins:
            # Create bins based on the data
            if series_clean.dtype in ['object', 'category']:
                # For categorical data, use unique values as bins
                unique_values = sorted(series_clean.unique())
                bin_edges = unique_values
                distribution = series_clean.value_counts(normalize=True).reindex(
                    unique_values, fill_value=0
                ).values
            else:
                # For numeric data, create quantile-based bins
                try:
                    _, bin_edges = np.histogram(series_clean, bins=self.bins)
                    # Adjust edges to include all data
                    bin_edges[0] = series_clean.min() - 1e-6
                    bin_edges[-1] = series_clean.max() + 1e-6
                    
                    # Calculate distribution
                    counts, _ = np.histogram(series_clean, bins=bin_edges)
                    distribution = counts / len(series_clean)
                    
                    # Ensure minimum bin size
                    if np.any(distribution < self.min_bin_size) and len(distribution) > 2:
                        # Merge small bins
                        distribution = self._merge_small_bins(distribution, self.min_bin_size)
                        
                except Exception as e:
                    logger.warning(f"Failed to create bins for {column}: {e}. Using equal-width bins.")
                    bin_edges = np.linspace(series_clean.min(), series_clean.max(), self.bins + 1)
                    counts, _ = np.histogram(series_clean, bins=bin_edges)
                    distribution = counts / len(series_clean)
        
        else:
            # Use existing bins
            if column not in self.bin_edges_:
                raise StabilityError(f"No fitted bins found for column {column}")
            
            bin_edges = self.bin_edges_[column]
            
            if isinstance(bin_edges, list):
                # Categorical bins
                distribution = series_clean.value_counts(normalize=True).reindex(
                    bin_edges, fill_value=0
                ).values
            else:
                # Numeric bins
                counts, _ = np.histogram(series_clean, bins=bin_edges)
                distribution = counts / len(series_clean)
        
        # Add small epsilon to avoid log(0) in PSI calculation
        distribution = np.maximum(distribution, 1e-10)
        
        return bin_edges, distribution
    
    def _merge_small_bins(self, distribution: np.ndarray, min_size: float) -> np.ndarray:
        """Merge bins that are smaller than minimum size."""
        merged_dist = distribution.copy()
        
        i = 0
        while i < len(merged_dist) - 1:
            if merged_dist[i] < min_size:
                # Merge with next bin
                merged_dist[i + 1] += merged_dist[i]
                merged_dist = np.delete(merged_dist, i)
            else:
                i += 1
        
        return merged_dist
    
    def _calculate_psi(self, reference_dist: np.ndarray, current_dist: np.ndarray) -> float:
        """Calculate Population Stability Index."""
        # Ensure distributions have same length
        min_len = min(len(reference_dist), len(current_dist))
        ref_dist = reference_dist[:min_len]
        curr_dist = current_dist[:min_len]
        
        # Ensure no zeros
        ref_dist = np.maximum(ref_dist, 1e-10)
        curr_dist = np.maximum(curr_dist, 1e-10)
        
        # Calculate PSI
        psi = np.sum((curr_dist - ref_dist) * np.log(curr_dist / ref_dist))
        
        return psi
    
    def _interpret_psi(self, psi_value: float) -> str:
        """Interpret PSI value."""
        if psi_value < self.PSI_THRESHOLDS['stable']:
            return 'stable'
        elif psi_value < self.PSI_THRESHOLDS['slight_shift']:
            return 'slight_shift'
        else:
            return 'significant_shift'
    
    def _create_psi_summary_plot(self, output_path: Path, figsize: Tuple[int, int]) -> Optional[str]:
        """Create PSI summary plot."""
        try:
            # Prepare data for plotting
            plot_data = []
            for dataset_name, results in self.stability_results_.items():
                for column, result in results.items():
                    if 'psi' in result and not np.isnan(result['psi']):
                        plot_data.append({
                            'dataset': dataset_name,
                            'variable': column,
                            'psi': result['psi'],
                            'status': result['stability_status']
                        })
            
            if not plot_data:
                return None
            
            df_plot = pd.DataFrame(plot_data)
            
            # Create plot
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=figsize)
            
            # PSI distribution
            ax1.hist(df_plot['psi'], bins=20, alpha=0.7, edgecolor='black')
            ax1.axvline(self.PSI_THRESHOLDS['stable'], color='green', linestyle='--', 
                       label=f'Stable threshold ({self.PSI_THRESHOLDS["stable"]})')
            ax1.axvline(self.PSI_THRESHOLDS['slight_shift'], color='orange', linestyle='--',
                       label=f'Slight shift threshold ({self.PSI_THRESHOLDS["slight_shift"]})')
            ax1.set_xlabel('PSI Value')
            ax1.set_ylabel('Frequency')
            ax1.set_title('Distribution of PSI Values')
            ax1.legend()
            ax1.grid(True, alpha=0.3)
            
            # Status counts
            status_counts = df_plot['status'].value_counts()
            colors = {'stable': 'green', 'slight_shift': 'orange', 'significant_shift': 'red'}
            bar_colors = [colors.get(status, 'gray') for status in status_counts.index]
            
            ax2.bar(status_counts.index, status_counts.values, color=bar_colors, alpha=0.7)
            ax2.set_xlabel('Stability Status')
            ax2.set_ylabel('Count')
            ax2.set_title('Variables by Stability Status')
            ax2.tick_params(axis='x', rotation=45)
            
            plt.tight_layout()
            
            plot_file = output_path / 'psi_summary.png'
            plt.savefig(plot_file, dpi=300, bbox_inches='tight')
            plt.close()
            
            return str(plot_file)
            
        except Exception as e:
            logger.error(f"Failed to create PSI summary plot: {e}")
            return None
    
    def _create_distribution_plot(self, column: str, output_path: Path, 
                                figsize: Tuple[int, int]) -> Optional[str]:
        """Create distribution comparison plot for a column."""
        try:
            fig, ax = plt.subplots(figsize=figsize)
            
            # Plot reference distribution
            ref_dist = self.reference_distributions_[column]
            x_positions = range(len(ref_dist))
            
            width = 0.35
            ax.bar([x - width/2 for x in x_positions], ref_dist, width, 
                  label='Reference', alpha=0.7, color='blue')
            
            # Plot current distributions
            colors = ['red', 'green', 'orange', 'purple', 'brown']
            color_idx = 0
            
            for dataset_name, results in self.stability_results_.items():
                if column in results and 'current_distribution' in results[column]:
                    curr_dist = results[column]['current_distribution']
                    color = colors[color_idx % len(colors)]
                    
                    ax.bar([x + width/2 for x in x_positions], curr_dist, width,
                          label=f'{dataset_name} (PSI: {results[column]["psi"]:.3f})',
                          alpha=0.7, color=color)
                    
                    color_idx += 1
            
            ax.set_xlabel('Bins')
            ax.set_ylabel('Frequency')
            ax.set_title(f'Distribution Comparison: {column}')
            ax.legend()
            ax.grid(True, alpha=0.3)
            
            plt.tight_layout()
            
            plot_file = output_path / f'distribution_{column}.png'
            plt.savefig(plot_file, dpi=300, bbox_inches='tight')
            plt.close()
            
            return str(plot_file)
            
        except Exception as e:
            logger.error(f"Failed to create distribution plot for {column}: {e}")
            return None
    
    def _create_psi_heatmap(self, output_path: Path, figsize: Tuple[int, int]) -> Optional[str]:
        """Create PSI heatmap across datasets and variables."""
        try:
            # Prepare data for heatmap
            heatmap_data = {}
            for dataset_name, results in self.stability_results_.items():
                heatmap_data[dataset_name] = {}
                for column, result in results.items():
                    if 'psi' in result:
                        heatmap_data[dataset_name][column] = result['psi']
            
            if not heatmap_data:
                return None
            
            # Convert to DataFrame
            heatmap_df = pd.DataFrame(heatmap_data).T
            
            # Create heatmap
            fig, ax = plt.subplots(figsize=figsize)
            
            sns.heatmap(heatmap_df, annot=True, fmt='.3f', cmap='RdYlGn_r',
                       center=self.PSI_THRESHOLDS['stable'], ax=ax,
                       cbar_kws={'label': 'PSI Value'})
            
            ax.set_title('PSI Heatmap: Datasets vs Variables')
            ax.set_xlabel('Variables')
            ax.set_ylabel('Datasets')
            
            plt.tight_layout()
            
            plot_file = output_path / 'psi_heatmap.png'
            plt.savefig(plot_file, dpi=300, bbox_inches='tight')
            plt.close()
            
            return str(plot_file)
            
        except Exception as e:
            logger.error(f"Failed to create PSI heatmap: {e}")
            return None
    
    def _create_summary_sheet(self, wb: Workbook) -> None:
        """Create summary sheet in Excel workbook."""
        ws = wb.create_sheet("Summary")
        
        # Headers
        headers = ['Dataset', 'Variable', 'PSI', 'Status', 'Interpretation']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Data
        row = 2
        for dataset_name, results in self.stability_results_.items():
            for column, result in results.items():
                if 'error' in result:
                    continue
                
                ws.cell(row=row, column=1, value=dataset_name)
                ws.cell(row=row, column=2, value=column)
                ws.cell(row=row, column=3, value=result.get('psi', 'N/A'))
                
                status = result.get('stability_status', 'unknown')
                status_cell = ws.cell(row=row, column=4, value=status)
                
                # Color coding
                if status == 'stable':
                    status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif status == 'slight_shift':
                    status_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                elif status == 'significant_shift':
                    status_cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                
                # Interpretation
                interpretation = self._get_psi_interpretation(result.get('psi', 0))
                ws.cell(row=row, column=5, value=interpretation)
                
                row += 1
    
    def _create_dataset_sheet(self, wb: Workbook, dataset_name: str) -> None:
        """Create detailed sheet for a specific dataset."""
        ws = wb.create_sheet(f"Details_{dataset_name}")
        
        results = self.stability_results_[dataset_name]
        
        # Add dataset information
        ws.cell(row=1, column=1, value=f"Dataset: {dataset_name}").font = Font(size=14, bold=True)
        
        # Statistics summary
        psi_values = [r.get('psi', np.nan) for r in results.values() if 'psi' in r]
        valid_psi = [p for p in psi_values if not np.isnan(p)]
        
        if valid_psi:
            ws.cell(row=3, column=1, value="Summary Statistics:")
            ws.cell(row=4, column=1, value="Total Variables:")
            ws.cell(row=4, column=2, value=len(valid_psi))
            ws.cell(row=5, column=1, value="Mean PSI:")
            ws.cell(row=5, column=2, value=np.mean(valid_psi))
            ws.cell(row=6, column=1, value="Max PSI:")
            ws.cell(row=6, column=2, value=np.max(valid_psi))
            ws.cell(row=7, column=1, value="Stable Variables:")
            ws.cell(row=7, column=2, value=sum(1 for p in valid_psi if p < self.PSI_THRESHOLDS['stable']))
        
        # Detailed results table
        start_row = 10
        headers = ['Variable', 'PSI', 'Status', 'Bins', 'Reference Mean', 'Current Mean']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True)
        
        row = start_row + 1
        for column, result in results.items():
            if 'error' in result:
                continue
            
            ws.cell(row=row, column=1, value=column)
            ws.cell(row=row, column=2, value=result.get('psi', 'N/A'))
            ws.cell(row=row, column=3, value=result.get('stability_status', 'unknown'))
            
            # Additional details if available
            if 'bin_edges' in result:
                ws.cell(row=row, column=4, value=len(result['bin_edges']))
            
            row += 1
    
    def _create_distributions_sheet(self, wb: Workbook) -> None:
        """Create distributions comparison sheet."""
        ws = wb.create_sheet("Distributions")
        
        # This would contain detailed distribution data
        # Implementation would depend on specific requirements
        ws.cell(row=1, column=1, value="Distribution Data").font = Font(size=14, bold=True)
        ws.cell(row=2, column=1, value="Detailed distribution data for each variable...")
    
    def _add_excel_charts(self, wb: Workbook) -> None:
        """Add charts to Excel workbook."""
        # Implementation for adding charts would go here
        # This would create bar charts, histograms, etc.
        pass
    
    def _get_psi_interpretation(self, psi_value: float) -> str:
        """Get detailed interpretation of PSI value."""
        if np.isnan(psi_value):
            return "No data"
        elif psi_value < 0.1:
            return "No significant change"
        elif psi_value < 0.2:
            return "Slight population shift, monitor"
        else:
            return "Significant population shift, investigate"